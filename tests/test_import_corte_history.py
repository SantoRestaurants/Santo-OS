from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from scripts.import_corte_history import parse_workbook


def test_history_import_uses_named_venta_bruta_column(tmp_path: Path) -> None:
    path = tmp_path / "ingresos.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.append(["AMEX", "Fecha", "Total Bruto", "PayPal", "Venta Bruta", "Uber Eats"])
    ws.append([100, datetime(2026, 6, 30), 900, 20, 875, 30])
    workbook.save(path)

    records = parse_workbook(path, "restaurant-id")

    assert len(records) == 1
    assert records[0]["venta_bruta"] == 875.0
    assert records[0]["total_bruto"] == 900.0
    assert records[0]["paypal"] == 20.0
    assert records[0]["uber_eats"] == 30.0


def test_layered_ingresos_header_rebases_stale_dates_from_filename(tmp_path: Path) -> None:
    path = tmp_path / "01. Santo_Ingresos Enero 2026.xlsx"
    workbook = Workbook()
    ws = workbook.active
    ws.cell(2, 3, "VENTA BRUTA")
    ws.cell(2, 15, "Total Bruto")
    ws.cell(2, 18, "Propinas")
    ws.cell(2, 20, "Venta Bruta")
    for column, label in {3: "Amex", 4: "Debito", 5: "Credito", 6: "EFECTIVO", 7: "TOTAL", 8: "PAYPAL", 10: "UBEREATS", 12: "RAPPI", 14: "TOTAL"}.items():
        ws.cell(4, column, label)
    ws.cell(5, 2, datetime(2024, 1, 1))
    ws.cell(5, 3, 100)
    ws.cell(5, 7, 100)
    ws.cell(5, 15, 130)
    ws.cell(5, 18, 10)
    ws.cell(5, 20, 120)
    ws.cell(6, 2, datetime(2024, 2, 1))
    ws.cell(6, 20, 999)
    workbook.save(path)

    records = parse_workbook(path, "restaurant-id")

    assert records[0]["business_date"] == "2026-01-01"
    assert records[0]["total"] == 100
    assert records[0]["total_bruto"] == 130
    assert records[0]["venta_bruta"] == 120
    assert len(records) == 1
