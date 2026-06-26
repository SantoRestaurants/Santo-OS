"""Parse Santo forecast Excel files and output JSON for Supabase import."""
import glob
import json
import os
import sys
from datetime import datetime

import openpyxl


def normalize_header(val: str) -> str:
    v = val.strip().upper()
    v = v.replace("\u00c1", "A").replace("\u00c9", "E").replace("\u00cd", "I").replace("\u00d3", "O").replace("\u00da", "U")
    v = v.replace("\u00d1", "N")
    return v


def parse_forecast_file(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Hoja1"]

    unit = "SANTO"
    header_row = None

    for row_idx in range(1, ws.max_row + 1):
        row_vals = [str(ws.cell(row_idx, c).value or "").strip() for c in range(1, 15)]
        norm = " ".join([normalize_header(v) for v in row_vals])

        # Detect unit
        for v in row_vals:
            vu = v.upper()
            if vu in ("SANTO", "EL_FAROLITO", "LA_GLORIETA") and row_idx <= 5:
                unit = vu

        if "FECHA" in norm and "META" in norm:
            header_row = row_idx
            break

    if header_row is None:
        raise ValueError(f"No header row found")

    # Map columns from header
    col_map = {}
    for c in range(1, 15):
        h = normalize_header(str(ws.cell(header_row, c).value or ""))
        if "DIA" in h and "FECHA" not in h and "META" not in h:
            col_map["dia"] = c
        elif "FECHA" in h:
            col_map["fecha"] = c
        elif "META" in h and "VTA" in h:
            col_map["meta"] = c
        elif "VENTA" in h and "REAL" in h:
            col_map["venta"] = c
        elif "DIFERENCIA" in h:
            col_map["diff"] = c

    if "fecha" not in col_map or "meta" not in col_map:
        raise ValueError(f"Missing required columns. Found: {list(col_map.keys())}")

    def cell_val(row, key):
        c = col_map.get(key)
        if c is None:
            return None
        return ws.cell(row, c).value

    dates_meta = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_fecha = cell_val(row_idx, "fecha")
        raw_meta = cell_val(row_idx, "meta")
        raw_venta = cell_val(row_idx, "venta")
        raw_dia = cell_val(row_idx, "dia")

        if raw_fecha is None or raw_fecha == "":
            break

        # Parse date
        if isinstance(raw_fecha, datetime):
            fecha_str = raw_fecha.strftime("%Y-%m-%d")
        elif isinstance(raw_fecha, str) and raw_fecha.strip():
            try:
                fecha_str = datetime.strptime(raw_fecha[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
            except ValueError:
                continue
        else:
            continue

        # Parse meta
        meta = 0.0
        if isinstance(raw_meta, (int, float)):
            meta = float(raw_meta)
        elif isinstance(raw_meta, str) and raw_meta.strip():
            try:
                meta = float(raw_meta.replace(",", "").replace("$", ""))
            except ValueError:
                meta = 0.0

        # Parse venta real
        venta = 0.0
        if isinstance(raw_venta, (int, float)):
            venta = float(raw_venta)
        elif isinstance(raw_venta, str) and raw_venta.strip():
            try:
                venta = float(raw_venta.replace(",", "").replace("$", ""))
            except ValueError:
                venta = 0.0

        dia_str = str(raw_dia or "")

        dates_meta.append({
            "dia": dia_str,
            "fecha": fecha_str,
            "meta_vta": round(meta, 2),
            "venta_real": round(venta, 2),
            "diferencia": round(venta - meta, 2),
        })

    if not dates_meta:
        raise ValueError("No data rows found")

    month = dates_meta[0]["fecha"][:7]
    total_meta = sum(x["meta_vta"] for x in dates_meta)

    vta_por_dia = [
        {
            "dia": x["dia"],
            "fecha": x["fecha"],
            "meta_vta": x["meta_vta"],
            "venta_real": x["venta_real"],
            "diferencia": x["diferencia"],
        }
        for x in dates_meta
    ]

    return {
        "filename": os.path.basename(filepath),
        "unit": unit,
        "restaurant_key": unit,
        "month": month,
        "total_meta": round(total_meta, 2),
        "days": len(dates_meta),
        "vta_por_dia": vta_por_dia,
    }


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\dchac\OneDrive\Desktop\santo"
    files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    if not files:
        print(f"No Excel files found in {folder}")
        return

    results = []
    for f in files:
        try:
            data = parse_forecast_file(f)
            results.append(data)
            print(f"{data['filename']:45s} {data['unit']:12s} {data['month']:8s} {data['days']:2d}d  META MXN {data['total_meta']:>14,.2f}")
        except Exception as e:
            print(f"ERROR {os.path.basename(f)}: {e}")

    out_path = os.path.join(folder, "forecasts_parsed.json")
    with open(out_path, "w", encoding="utf-8") as fp:
        json.dump(results, fp, indent=2, ensure_ascii=False)
    print(f"\nWrote {len(results)} months to {out_path}")


if __name__ == "__main__":
    main()
