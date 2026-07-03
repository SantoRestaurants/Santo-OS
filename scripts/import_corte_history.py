"""Import historical Ingresos workbooks into ``corte_daily_records``.

Dry-run is the default. Pass ``--write`` only after reviewing the generated
JSON report. The import is idempotent on restaurant and business date.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import unicodedata
import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from workflows.corte_santo.daily_record import PARSER_VERSION, build_daily_record


HEADER_ALIASES = {
    "fecha": {"fecha"},
    "amex": {"amex"},
    "debito": {"debito", "t debito"},
    "credito": {"credito", "t credito"},
    "efectivo": {"efectivo", "efectivo sistema"},
    "transferencia": {"transferencia"},
    "total": {"total"},
    "paypal": {"paypal"},
    "uber": {"uber", "uber eats", "ubereats"},
    "rappi": {"rappi"},
    "propinas": {"propinas", "propina"},
    "venta_bruta": {"venta bruta"},
    "total_bruto": {"total bruto"},
}

SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5,
    "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9,
    "setiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )
    return " ".join(text.replace("_", " ").split())


def find_header(ws: Any) -> tuple[int, dict[str, int]]:
    max_row = ws.max_row or 100
    max_column = ws.max_column or 50
    for row_number in range(1, min(max_row, 30) + 1):
        normalized = {column: normalize(ws.cell(row_number, column).value) for column in range(1, max_column + 1)}
        mapping: dict[str, int] = {}
        for key, aliases in HEADER_ALIASES.items():
            match = next((column for column, value in normalized.items() if value in aliases), None)
            if match is not None:
                mapping[key] = match
        if {"fecha", "venta_bruta"}.issubset(mapping):
            return row_number, mapping
    raise ValueError("required_headers_not_found:fecha,venta_bruta")


def find_ingresos_layout(ws: Any) -> tuple[int, dict[str, int]]:
    """Read the real layered Ingresos header (summary row 2, channels row 4)."""
    max_row = ws.max_row or 100
    max_column = ws.max_column or 50
    data_start = None
    date_column = None
    for row_number in range(1, min(max_row, 30) + 1):
        for column in range(1, max_column + 1):
            if parse_date(ws.cell(row_number, column).value):
                data_start = row_number
                date_column = column
                break
        if data_start:
            break
    if data_start is None or date_column is None:
        raise ValueError("daily_date_column_not_found")

    columns: dict[str, int] = {"fecha": date_column}
    for column in range(1, max_column + 1):
        labels = [normalize(ws.cell(row, column).value) for row in range(1, data_start)]
        channel_label = labels[-1] if labels else ""
        for key in ("amex", "debito", "credito", "efectivo", "paypal", "uber", "rappi"):
            if channel_label in HEADER_ALIASES[key] and key not in columns:
                columns[key] = column

    summary_labels = {
        column: normalize(ws.cell(2, column).value)
        for column in range(1, max_column + 1)
    }
    for key in ("total_bruto", "propinas", "venta_bruta"):
        matches = [column for column, label in summary_labels.items() if label in HEADER_ALIASES[key]]
        if matches:
            columns[key] = matches[-1]

    total_candidates = [
        column for column in range(1, max_column + 1)
        if normalize(ws.cell(data_start - 1, column).value) == "total"
    ]
    if total_candidates:
        columns["total"] = total_candidates[0]
    if "venta_bruta" not in columns:
        raise ValueError("required_header_not_found:venta_bruta")
    return data_start - 1, columns


def month_from_filename(path: Path) -> tuple[int, int] | None:
    normalized = normalize(path.stem)
    year_match = re.search(r"\b(20\d{2})\b", normalized)
    month = next((number for name, number in SPANISH_MONTHS.items() if name in normalized), None)
    if not year_match or month is None:
        return None
    return int(year_match.group(1)), month


def parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()[:10]
        try:
            return datetime.strptime(text, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return None
    return None


def parse_workbook(path: Path, restaurant_id: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    ws = workbook.active
    try:
        header_row, columns = find_header(ws)
    except ValueError:
        header_row, columns = find_ingresos_layout(ws)
    source_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    forced_period = month_from_filename(path)
    first_source_date = parse_date(ws.cell(header_row + 1, columns["fecha"]).value)
    source_period = first_source_date[:7] if first_source_date else None
    records: list[dict[str, Any]] = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        business_date = parse_date(ws.cell(row_number, columns["fecha"]).value)
        if not business_date:
            continue
        if source_period and business_date[:7] != source_period:
            continue
        if forced_period:
            forced_year, forced_month = forced_period
            if business_date[:7] != f"{forced_year:04d}-{forced_month:02d}":
                day = datetime.strptime(business_date, "%Y-%m-%d").day
                if day > monthrange(forced_year, forced_month)[1]:
                    continue
                business_date = date(forced_year, forced_month, day).isoformat()
        values = {
            key: ws.cell(row_number, column).value
            for key, column in columns.items()
            if key not in {"fecha", "venta_bruta", "total_bruto", "total"}
        }
        record = build_daily_record(
            restaurant_id=restaurant_id,
            business_date=business_date,
            income_register=values,
            venta_bruta=ws.cell(row_number, columns["venta_bruta"]).value,
            total=ws.cell(row_number, columns["total"]).value if "total" in columns else None,
            total_bruto=ws.cell(row_number, columns["total_bruto"]).value if "total_bruto" in columns else None,
            source_kind="historical_import",
            source_filename=path.name,
            source_sheet=ws.title,
            source_row=row_number,
            source_hash=source_hash,
            parser_version=PARSER_VERSION,
        )
        if record["venta_bruta"] is not None:
            records.append(record)

    workbook.close()
    return records


def resolve_restaurant_id(client: httpx.Client, restaurant_key: str) -> str:
    response = client.get(
        "/rest/v1/restaurants",
        params={"restaurant_key": f"eq.{restaurant_key}", "select": "id", "limit": "1"},
    )
    response.raise_for_status()
    rows = response.json()
    if not rows:
        raise RuntimeError(f"restaurant_not_found:{restaurant_key}")
    return str(rows[0]["id"])


def upsert_records(client: httpx.Client, records: list[dict[str, Any]]) -> None:
    for start in range(0, len(records), 500):
        response = client.post(
            "/rest/v1/corte_daily_records",
            params={"on_conflict": "restaurant_id,business_date"},
            headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
            json=records[start:start + 500],
        )
        response.raise_for_status()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("paths", nargs="+", type=Path, help="Workbook files or folders")
    parser.add_argument("--restaurant-key", default="default_restaurant_confirm")
    parser.add_argument("--restaurant-id", help="Allows offline dry-run without Supabase lookup")
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--report", type=Path, default=Path("artifacts/corte_history_import.json"))
    args = parser.parse_args()

    files: list[Path] = []
    for path in args.paths:
        files.extend(sorted(path.glob("*.xlsx")) if path.is_dir() else [path])

    url = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_SECRET_KEY")
    restaurant_id = args.restaurant_id
    client = None
    if args.write or not restaurant_id:
        if not url or not key:
            raise RuntimeError("Supabase URL and service key are required")
        client = httpx.Client(
            base_url=url.rstrip("/"),
            headers={"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"},
            timeout=60,
        )
        restaurant_id = resolve_restaurant_id(client, args.restaurant_key)

    assert restaurant_id
    records = [record for path in files for record in parse_workbook(path, restaurant_id)]
    report = {
        "mode": "write" if args.write else "dry_run",
        "parser_version": PARSER_VERSION,
        "files": [str(path) for path in files],
        "record_count": len(records),
        "first_date": min((row["business_date"] for row in records), default=None),
        "last_date": max((row["business_date"] for row in records), default=None),
        "records": records,
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    if args.write:
        assert client
        upsert_records(client, records)
    if client:
        client.close()
    print(json.dumps({key: value for key, value in report.items() if key != "records"}, indent=2))


if __name__ == "__main__":
    main()
