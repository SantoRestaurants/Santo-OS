"""
Corte Santo Excel parser (Option B automation).

Extracts the "Cierre Ter/Pla" (terminals/platforms = real) and "Cierre Sistema"
(Wansoft POS) blocks from the client's "SANTO CORTE ..." workbook and turns them
into the structured `cierre_terminal` / `cierre_sistema` payload the workflow
reconciles.

Design constraints (P0):

- Nothing about the layout is hardcoded as a business assumption. Anchor labels,
  the column-label -> reconciliation-group map and the row labels are driven by
  config (`excel_layout`). A confirmed default is shipped for the SANTO unit.
- Any column label the parser cannot confidently map to a group is reported as a
  warning so the workflow can return `requires_review` instead of silently
  dropping or miscounting money. Uncertainty never becomes `completed`.
- Read-only, data_only parsing (computed values, no macros, no formula eval).
"""

from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Any

try:  # openpyxl is an optional dependency; absence is handled by the caller.
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - exercised only when dependency missing
    load_workbook = None  # type: ignore[assignment]


DEFAULT_LAYOUT: dict[str, Any] = {
    "terminal_anchor": "Cierre Ter/Pla",
    "sistema_anchor": "Cierre Sistema",
    "consumo_label": "Consumo",
    "propina_label": "Propina",
    # In the client's corte template these columns repeat the cash amount on
    # the row labelled "Propina", but that repeated value is the comparison
    # amount/global, not a cash tip. Counting it as propina doubles cash.
    "propina_is_global_labels": ["efectivo real", "efectivo sistema"],
    # The client template keeps Transferencia/Uber/Rappi system comparison
    # values in a supplemental block below the main Cierre Sistema table.
    "supplemental_system_anchor": "Total Sistema",
    "supplemental_header_anchor": "Total Real",
    "supplemental_value_row_offset": 1,
    # Maps a normalized column header to a reconciliation group, or to null to
    # explicitly ignore aggregate/total columns so they are not double counted.
    "column_label_map": {
        "amex": "amex",
        "bancos": "bancos",
        "t debito": "bancos",
        "t credito": "bancos",
        "total bancos": None,
        "efectivo": "efectivo",
        "efectivo real": "efectivo",
        "efectivo sistema": "efectivo",
        "transferencia": "transferencia",
        "uber eats": "plataformas",
        "rappi": "plataformas",
        "paypal": "paypal",
        "global": None,
        "total": None,
    },
    "max_columns": 12,
}


def _normalize(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = "".join(
        ch for ch in unicodedata.normalize("NFD", value) if unicodedata.category(ch) != "Mn"
    )
    return " ".join(value.split())


def _to_amount(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "").replace(" ", "")
    if text in ("", "-"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        amount = float(text)
    except ValueError:
        return 0.0
    return -amount if negative else amount


def _find_anchor(rows: list[list[Any]], anchor_norm: str) -> tuple[int, int] | None:
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            if _normalize(cell) == anchor_norm:
                return (r, c)
    return None


def _find_row_label(
    rows: list[list[Any]], start_row: int, anchor_col: int, label_norm: str, window: int = 4
) -> int | None:
    for r in range(start_row, min(start_row + window + 1, len(rows))):
        if _normalize(rows[r][anchor_col] if anchor_col < len(rows[r]) else None) == label_norm:
            return r
    return None


def _parse_block(
    rows: list[list[Any]],
    anchor_norm: str,
    layout: dict[str, Any],
) -> tuple[dict[str, dict[str, float]], list[str]]:
    warnings: list[str] = []
    anchor = _find_anchor(rows, anchor_norm)
    if anchor is None:
        return {}, [f"anchor_not_found:{anchor_norm}"]

    anchor_row, anchor_col = anchor
    column_label_map = layout.get("column_label_map", {})
    propina_is_global_labels = {
        _normalize(label) for label in layout.get("propina_is_global_labels", [])
    }
    max_columns = int(layout.get("max_columns", 12))

    # Header labels live in the same row as the anchor, to the right of it.
    header_cols: list[tuple[int, str]] = []
    header_row = rows[anchor_row]
    for c in range(anchor_col + 1, min(anchor_col + 1 + max_columns, len(header_row))):
        label = _normalize(header_row[c])
        if label:
            header_cols.append((c, label))

    consumo_row = _find_row_label(rows, anchor_row + 1, anchor_col, _normalize(layout["consumo_label"]))
    propina_row = _find_row_label(rows, anchor_row + 1, anchor_col, _normalize(layout["propina_label"]))
    if consumo_row is None or propina_row is None:
        warnings.append(f"rows_not_found:{anchor_norm}")
        return {}, warnings

    groups: dict[str, dict[str, float]] = {}
    for col, label in header_cols:
        if label not in column_label_map:
            warnings.append(f"unmapped_column:{anchor_norm}:{label}")
            continue
        group = column_label_map[label]
        if group is None:  # explicitly ignored aggregate/total column
            continue
        consumo = _to_amount(rows[consumo_row][col] if col < len(rows[consumo_row]) else None)
        propina = _to_amount(rows[propina_row][col] if col < len(rows[propina_row]) else None)
        if label in propina_is_global_labels:
            propina = 0.0
        bucket = groups.setdefault(group, {"consumo": 0.0, "propina": 0.0})
        bucket["consumo"] = round(bucket["consumo"] + consumo, 2)
        bucket["propina"] = round(bucket["propina"] + propina, 2)

    return groups, warnings


def parse_corte_workbook(rows: list[list[Any]], config: dict[str, Any] | None = None) -> dict[str, Any]:
    """Parse pre-loaded worksheet rows into terminal/sistema group dicts."""
    config = config or {}
    layout = {**DEFAULT_LAYOUT, **(config.get("excel_layout") or {})}

    terminal, terminal_warnings = _parse_block(rows, _normalize(layout["terminal_anchor"]), layout)
    sistema, sistema_warnings = _parse_block(rows, _normalize(layout["sistema_anchor"]), layout)

    supplemental_warnings: list[str] = []
    supplemental_anchor = _find_anchor(rows, _normalize(layout["supplemental_system_anchor"]))
    supplemental_headers = _find_anchor(rows, _normalize(layout["supplemental_header_anchor"]))
    if supplemental_anchor and supplemental_headers:
        value_row = supplemental_anchor[0] + int(layout.get("supplemental_value_row_offset", 1))
        header_row = rows[supplemental_headers[0]]
        column_label_map = layout.get("column_label_map", {})
        for col in range(supplemental_headers[1] + 1, len(header_row)):
            label = _normalize(header_row[col])
            if not label:
                continue
            if label not in column_label_map:
                supplemental_warnings.append(f"unmapped_supplemental_column:{label}")
                continue
            group = column_label_map[label]
            if group is None:
                continue
            value = _to_amount(
                rows[value_row][col]
                if value_row < len(rows) and col < len(rows[value_row])
                else None
            )
            bucket = sistema.setdefault(group, {"consumo": 0.0, "propina": 0.0})
            bucket["consumo"] = round(bucket["consumo"] + value, 2)

    warnings = terminal_warnings + sistema_warnings + supplemental_warnings
    return {
        "cierre_terminal": terminal,
        "cierre_sistema": sistema,
        "warnings": warnings,
    }


def parse_corte_excel(source_path: str, config: dict[str, Any] | None = None) -> dict[str, Any]:
    """Load an .xlsx corte workbook and parse its terminal/sistema blocks."""
    if load_workbook is None:
        return {
            "cierre_terminal": {},
            "cierre_sistema": {},
            "warnings": ["openpyxl_not_available"],
        }

    path = Path(source_path)
    if not path.exists():
        return {
            "cierre_terminal": {},
            "cierre_sistema": {},
            "warnings": [f"file_not_found:{source_path}"],
        }

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return parse_corte_workbook(rows, config)
