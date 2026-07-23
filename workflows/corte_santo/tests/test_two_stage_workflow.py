from __future__ import annotations

import importlib.util
from datetime import datetime
from pathlib import Path

from openpyxl.comments import Comment
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]


def _load(name: str):
    spec = importlib.util.spec_from_file_location(name, ROOT / f"{name}.py")
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


writer = _load("workbook_writer")
bank = _load("bank_reconciliation")
bank_parser = _load("bank_statement_parser")
pipeline = _load("two_stage_pipeline")
runtime = _load("runtime")


def _ingresos(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["B5"] = datetime(2026, 6, 4)
    wb.save(path)


def _ingresos_with_formula_dates(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["B5"] = datetime(2026, 6, 1)
    ws["B6"] = "=+B5+1"
    ws["B7"] = "=+B6+1"
    ws["B8"] = "=+B7+1"
    wb.save(path)


def _forecast(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["C4"] = datetime(2026, 6, 4)
    ws["C38"] = "TOTAL MES"
    wb.save(path)


def _forecast_stale_projection_month(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    for day in range(1, 31):
        row = day + 3
        ws.cell(row, 3).value = datetime(2026, 5, day)
        ws.cell(row, 4).value = day * 1000
    ws["C38"] = "TOTAL MES"
    wb.save(path)


VALUES = {
    "amex": 9909.45,
    "debito": 5130.25,
    "credito": 52061.9,
    "efectivo": 5138.5,
    "paypal": 0,
    "uber": 3525,
    "rappi": 0,
    "propinas": 6582.6,
}
INGRESOS_LAYOUT = {
    "date_column": 2,
    "columns": {"amex": "C", "debito": "D", "credito": "E", "efectivo": "F", "paypal": "H", "uber": "J", "rappi": "L", "propinas": "R"},
    "stage_colors": {"corte_loaded": writer.YELLOW, "bank_validated": writer.BLUE},
}
FORECAST_LAYOUT = {
    "date_column": 3,
    "venta_real_column": "E",
    "total_month_label_column": "C",
    "total_month_label": "TOTAL MES",
    "subtotal_columns": ["D", "E", "F"],
    "data_start_row": 4,
    "data_end_row": 34,
}


def test_ingresos_moves_from_yellow_to_blue(tmp_path: Path) -> None:
    source = tmp_path / "ingresos.xlsx"
    yellow = tmp_path / "yellow.xlsx"
    blue = tmp_path / "blue.xlsx"
    _ingresos(source)

    first = writer.write_ingresos(str(source), str(yellow), "2026-06-04", VALUES, dry_run=False, layout=INGRESOS_LAYOUT)
    second = writer.write_ingresos(
        str(yellow), str(blue), "2026-06-04", VALUES, stage="bank_validated", dry_run=False, layout=INGRESOS_LAYOUT
    )

    assert first["status"] == "written"
    assert second["status"] == "written"
    assert load_workbook(yellow).active["C5"].fill.fgColor.rgb == writer.YELLOW
    assert load_workbook(blue).active["C5"].fill.fgColor.rgb == writer.BLUE


def test_ingresos_finds_cached_formula_date_rows(tmp_path: Path) -> None:
    source = tmp_path / "ingresos-formulas.xlsx"
    output = tmp_path / "ingresos-output.xlsx"
    _ingresos_with_formula_dates(source)

    result = writer.write_ingresos(
        str(source), str(output), "2026-06-04", VALUES, dry_run=False, layout=INGRESOS_LAYOUT
    )

    assert result["status"] == "written"
    assert load_workbook(output).active["C8"].value == VALUES["amex"]


def test_ingresos_preserves_red_paypal_adjustment_comment(tmp_path: Path) -> None:
    source = tmp_path / "ingresos-paypal-comment.xlsx"
    output = tmp_path / "ingresos-paypal-output.xlsx"
    _ingresos(source)
    wb = load_workbook(source)
    ws = wb.active
    ws["H5"] = "=2395-2395"
    ws["H5"].fill = writer.PatternFill("solid", fgColor=writer.RED)
    ws["H5"].comment = Comment("Mov 89109 $2395 con propina de $359.25 del sr. Toro", "None")
    wb.save(source)
    wb.close()

    result = writer.write_ingresos(
        str(source),
        str(output),
        "2026-06-04",
        VALUES,
        stage="bank_validated",
        dry_run=False,
        layout=INGRESOS_LAYOUT,
    )

    assert result["status"] == "written"
    ws_out = load_workbook(output).active
    assert ws_out["H5"].value == "=2395-2395"
    assert ws_out["H5"].fill.fgColor.rgb == writer.RED
    assert "propina de $359.25" in ws_out["H5"].comment.text


def test_ingresos_writes_cxc_paypal_comment_when_note_is_supplied(tmp_path: Path) -> None:
    source = tmp_path / "ingresos-paypal-note.xlsx"
    output = tmp_path / "ingresos-paypal-note-output.xlsx"
    _ingresos(source)

    result = writer.write_ingresos(
        str(source),
        str(output),
        "2026-06-04",
        VALUES,
        stage="corte_loaded",
        dry_run=False,
        layout=INGRESOS_LAYOUT,
        cell_notes={
            "paypal": {
                "kind": "cxc",
                "amount": 6714.0,
                "formula": "=6714-6714",
                "comment": "CXC\nMOV 87745 $1,695\nTOTAL $6714\n======",
            }
        },
    )

    assert result["status"] == "written"
    ws_out = load_workbook(output).active
    assert ws_out["H5"].value == "=6714-6714"
    assert ws_out["H5"].fill.fgColor.rgb == writer.RED
    assert "MOV 87745" in ws_out["H5"].comment.text


def test_ingresos_updates_existing_paypal_note_when_note_is_supplied(tmp_path: Path) -> None:
    source = tmp_path / "ingresos-paypal-existing-note.xlsx"
    output = tmp_path / "ingresos-paypal-existing-note-output.xlsx"
    _ingresos(source)
    wb = load_workbook(source)
    ws = wb.active
    ws["H5"] = "=245-245"
    ws["H5"].fill = writer.PatternFill("solid", fgColor=writer.RED)
    ws["H5"].comment = Comment("old OCR note", "SantoOS")
    wb.save(source)
    wb.close()

    result = writer.write_ingresos(
        str(source),
        str(output),
        "2026-06-04",
        VALUES,
        stage="corte_loaded",
        dry_run=False,
        layout=INGRESOS_LAYOUT,
        cell_notes={
            "paypal": {
                "kind": "cxc",
                "amount": 245.0,
                "formula": "=245-245",
                "comment": "CXC\nPago en efectivo de CXC: $245.00\n======",
            }
        },
    )

    assert result["status"] == "written"
    ws_out = load_workbook(output, data_only=False).active
    assert ws_out["H5"].value == "=245-245"
    assert "Pago en efectivo de CXC" in ws_out["H5"].comment.text


def test_ingresos_writes_paypal_value_with_cxc_note(tmp_path: Path) -> None:
    source = tmp_path / "ingresos-paypal-value-note.xlsx"
    output = tmp_path / "ingresos-paypal-value-note-output.xlsx"
    _ingresos(source)
    values = {**VALUES, "paypal": 513.0}

    result = writer.write_ingresos(
        str(source),
        str(output),
        "2026-06-04",
        values,
        stage="corte_loaded",
        dry_run=False,
        layout=INGRESOS_LAYOUT,
        cell_notes={
            "paypal": {
                "kind": "cxc",
                "amount": 513.0,
                "formula": "=3078-2565",
                "comment": "CXC\nCXC MESERO MOV 89972 $245\nPAGO CXC MOV 87028 TRANSFERENCIA\n======",
            }
        },
    )

    assert result["status"] == "written"
    ws_out = load_workbook(output, data_only=False).active
    assert ws_out["H5"].value == "=3078-2565"
    assert ws_out["H5"].fill.fgColor.rgb == writer.YELLOW
    assert "MOV 89972 $245" in ws_out["H5"].comment.text


def test_forecast_write_updates_total_month_formula(tmp_path: Path) -> None:
    source = tmp_path / "forecast.xlsx"
    output = tmp_path / "forecast-output.xlsx"
    _forecast(source)

    result = writer.write_forecast(str(source), str(output), "2026-06-04", 75685.1, dry_run=False, layout=FORECAST_LAYOUT)

    ws = load_workbook(output, data_only=False).active
    assert result["status"] == "written"
    assert ws["E4"].value == 75685.1
    assert ws["E38"].value == "=+SUBTOTAL(9,E4:E34)"


def test_forecast_rebases_confirmed_projection_month(tmp_path: Path) -> None:
    source = tmp_path / "forecast-stale-month.xlsx"
    output = tmp_path / "forecast-june.xlsx"
    _forecast_stale_projection_month(source)
    layout = {**FORECAST_LAYOUT, "allow_projection_month_rebase": True}

    result = writer.write_forecast(
        str(source), str(output), "2026-06-04", 75685.1, dry_run=False, layout=layout
    )

    ws = load_workbook(output, data_only=False).active
    assert result["status"] == "written"
    assert ws["C4"].value == datetime(2026, 6, 1)
    assert ws["C33"].value == datetime(2026, 6, 30)
    assert ws["D7"].value == 4000
    assert ws["E7"].value == 75685.1


def test_amex_named_columns_are_parsed() -> None:
    result = bank.parse_amex_rows(
        [
            ["metadata"],
            ["Fecha de pago", "Monto del pago"],
            ["2026-06-05", "9,500.25"],
        ]
    )
    assert result["status"] == "ok"
    assert result["total_expected"] == 9500.25


def test_bank_matching_allows_legitimate_pending_collections() -> None:
    result = bank.reconcile_bank_stage(
        [{"channel": "amex", "amount": 100.0}, {"channel": "amex", "amount": 50.0}],
        {
            "status": "ok",
            "deposits": [{"source": "amex", "amount": 100.0}],
            "domiciled_expenses": [],
        },
        {
            "status": "ok",
            "payments": [{"amount": 100.0, "gross_amount": 100.0}]
        },
    )
    assert result["status"] == "bank_validated"
    assert result["pending_collections"]["amex"] == 50.0


def test_bank_matching_dedupes_equivalent_pending_amex_representations() -> None:
    result = bank.reconcile_bank_stage(
        [
            {
                "business_date": "2026-06-25",
                "source_date": "2026-06-25",
                "channel": "amex",
                "amount": 16277.36,
                "expected_deposit": 16277.36,
                "expected_payment_date": "2026-06-30",
            }
        ],
        {
            "status": "ok",
            "deposits": [],
            "domiciled_expenses": [],
        },
        {
            "status": "ok",
            "payments": [
                {
                    "amount": 16277.36,
                    "gross_amount": 16277.36,
                    "source_date": "2026-06-25",
                    "payment_date": "2026-06-30",
                }
            ],
        },
    )

    assert result["status"] == "bank_validated"
    assert result["pending_collections"]["amex"] == 16277.36
    assert len(result["pending_items"]) == 1


def test_bank_matching_groups_amex_payments_by_expected_payment_date() -> None:
    result = bank.reconcile_bank_stage(
        [
            {"channel": "amex", "amount": 18437.66, "source_date": "12/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 62832.78, "source_date": "13/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 37299.77, "source_date": "14/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 689.89, "source_date": "14/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 94.47, "source_date": "12/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 7467.54, "source_date": "14/6/2026", "expected_payment_date": "17/6/2026"},
            {"channel": "amex", "amount": 24309.8, "source_date": "14/6/2026", "expected_payment_date": "17/6/2026"},
        ],
        {
            "status": "ok",
            "deposits": [{"source": "amex", "amount": 151131.91, "operation_date": "17/06/2026"}],
            "domiciled_expenses": [],
        },
        {
            "status": "ok", 
            "payments": [
                {"amount": 18437.66, "gross_amount": 18437.66, "source_date": "12/6/2026", "payment_date": "17/6/2026"},
                {"amount": 62832.78, "gross_amount": 62832.78, "source_date": "13/6/2026", "payment_date": "17/6/2026"},
                {"amount": 37299.77, "gross_amount": 37299.77, "source_date": "14/6/2026", "payment_date": "17/6/2026"},
                {"amount": 689.89, "gross_amount": 689.89, "source_date": "14/6/2026", "payment_date": "17/6/2026"},
                {"amount": 94.47, "gross_amount": 94.47, "source_date": "12/6/2026", "payment_date": "17/6/2026"},
                {"amount": 7467.54, "gross_amount": 7467.54, "source_date": "14/6/2026", "payment_date": "17/6/2026"},
                {"amount": 24309.8, "gross_amount": 24309.8, "source_date": "14/6/2026", "payment_date": "17/6/2026"},
            ]
        },
    )
    assert result["status"] == "bank_validated"
    assert result["pending_collections"] == {}
    assert result["matches"][0]["expected_group"][0]["amount"] == 18437.66


def test_bank_matching_uses_corte_ledger_as_canonical_amex_source() -> None:
    result = bank.reconcile_bank_stage(
        [
            {"business_date": "2026-07-02", "channel": "amex", "amount": 1000.0, "source_date": "2026-07-02"},
            {"business_date": "2026-07-03", "channel": "amex", "amount": 2000.0, "source_date": "2026-07-03"},
        ],
        {
            "status": "ok",
            "deposits": [{"source": "amex", "amount": 960.0, "operation_date": "2026-07-07"}],
            "additional_expenses": [],
        },
        {
            "status": "ok",
            "payments": [
                {
                    "cargos": 1000.0,
                    "neto": 960.0,
                    "amount": 960.0,
                    "gross_amount": 1000.0,
                    "source_date": "2026-07-02",
                    "fecha_envio": "2026-07-02",
                    "payment_date": "2026-07-07",
                },
                {
                    "cargos": 9999.0,
                    "neto": 9700.0,
                    "amount": 9700.0,
                    "gross_amount": 9999.0,
                    "source_date": "2026-07-04",
                    "fecha_envio": "2026-07-04",
                    "payment_date": "2026-07-08",
                },
            ],
        },
    )

    assert result["pending_collections"] == {"amex": 2000.0}
    assert result["pending_items"] == [
        {
            "business_date": "2026-07-03",
            "channel": "amex",
            "amount": 2000.0,
            "source_date": "2026-07-03",
            "expected_deposit": 2000.0,
            "status": "pendiente_reporte_amex",
        }
    ]


def test_bank_matching_clears_banorte_group_and_keeps_platforms_separate() -> None:
    result = bank.reconcile_bank_stage(
        [
            {"business_date": "2026-07-06", "channel": "banorte", "amount": 100.0, "source_date": "2026-07-06"},
            {"business_date": "2026-07-06", "channel": "uber", "amount": 60.0, "source_date": "2026-07-06"},
            {"business_date": "2026-07-06", "channel": "rappi", "amount": 50.0, "source_date": "2026-07-06"},
        ],
        {
            "status": "ok",
            "deposits": [
                {"source": "banorte", "amount": 40.0, "operation_date": "2026-07-07"},
                {"source": "banorte", "amount": 60.0, "operation_date": "2026-07-07"},
                {"source": "uber", "amount": 60.0, "operation_date": "2026-07-07"},
            ],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
    )

    assert result["pending_collections"] == {"rappi": 50.0}
    assert [item["channel"] for item in result["pending_items"]] == ["rappi"]


def test_banorte_same_day_deposit_stays_pending_until_next_day() -> None:
    result = bank.reconcile_bank_stage(
        [{
            "business_date": "2026-07-16",
            "source_date": "2026-07-16",
            "channel": "banorte",
            "amount": 63444.0,
        }],
        {
            "status": "ok",
            "deposits": [{"source": "banorte", "amount": 63444.0, "operation_date": "16/07/2026"}],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
        settlement_rules={"banorte": {"mode": "fifo_partial"}},
    )

    assert result["pending_collections"] == {"banorte": 63444.0}
    assert result["matches"] == []


def test_banorte_next_batch_carries_prior_residual_and_leaves_latest_day_partial() -> None:
    result = bank.reconcile_bank_stage(
        [
            {
                "business_date": "2026-07-14",
                "source_date": "2026-07-14",
                "channel": "banorte",
                "amount": 23178.30,
                "original_amount": 69084.24,
                "settled_amount": 45905.94,
                "status": "parcialmente_depositado",
            },
            {
                "business_date": "2026-07-15",
                "source_date": "2026-07-15",
                "channel": "banorte",
                "amount": 72742.66,
            },
        ],
        {
            "status": "ok",
            "deposits": [
                {"source": "banorte", "amount": 37296.71, "operation_date": "16/07/2026"},
                {"source": "banorte", "amount": 25004.65, "operation_date": "16/07/2026"},
                {"source": "banorte", "amount": 11543.00, "operation_date": "16/07/2026"},
                {"source": "banorte", "amount": 2760.50, "operation_date": "16/07/2026"},
            ],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
        settlement_rules={"banorte": {"mode": "fifo_partial"}},
    )

    assert result["pending_collections"] == {"banorte": 19316.10}
    assert result["pending_items"] == [{
        "business_date": "2026-07-15",
        "source_date": "2026-07-15",
        "channel": "banorte",
        "amount": 19316.10,
        "expected_deposit": 19316.10,
        "original_amount": 72742.66,
        "settled_amount": 53426.56,
        "status": "parcialmente_depositado",
    }]


def test_cxc_bank_transfer_allocates_partially_to_receivable() -> None:
    result = bank.reconcile_bank_stage(
        [{
            "business_date": "2026-07-09",
            "source_date": "2026-07-09",
            "channel": "cxc",
            "amount": 3185.0,
            "receivable_id": "la-valisse",
            "receivable_key": "restaurant:manual:la-valisse",
        }],
        {
            "status": "ok",
            "deposits": [{"source": "cxc", "amount": 1410.0, "operation_date": "18/07/2026"}],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
        settlement_rules={"cxc": {"mode": "fifo_partial"}},
    )

    assert result["pending_collections"] == {"cxc": 1775.0}
    allocation = result["matches"][0]["allocations"][0]
    assert allocation["receivable_id"] == "la-valisse"
    assert allocation["amount"] == 1410.0


def test_pending_amex_with_payment_date_is_cleared_when_batch_deposit_arrives() -> None:
    result = bank.reconcile_bank_stage(
        [{
            "business_date": "2026-07-07",
            "source_date": "2026-07-07",
            "channel": "amex",
            "amount": 14594.26,
            "expected_deposit": 14594.26,
            "expected_payment_date": "2026-07-13",
            "status": "pendiente_reporte_amex",
        }],
        {
            "status": "ok",
            "deposits": [{"source": "amex", "amount": 14594.26, "operation_date": "13/07/2026"}],
            "additional_expenses": [],
        },
        {
            "status": "ok",
            "payments": [{
                "pago_num": "41909428",
                "cargos": 15102.30,
                "neto": 14594.26,
                "fecha_envio": "2026-07-07",
                "payment_date": "2026-07-13",
            }],
        },
    )

    assert result["pending_collections"] == {}
    assert result["pending_items"] == []
    assert result["batch_validation"][0]["status"] == "ok"
    assert result["matches"][0]["expected"]["amount"] == 15102.30


def test_banorte_later_deposits_apply_fifo_and_keep_only_partial_residual() -> None:
    result = bank.reconcile_bank_stage(
        [
            {"business_date": "2026-07-09", "source_date": "2026-07-09", "channel": "banorte", "amount": 72864.86},
            {"business_date": "2026-07-10", "source_date": "2026-07-10", "channel": "banorte", "amount": 114083.55},
            {"business_date": "2026-07-11", "source_date": "2026-07-11", "channel": "banorte", "amount": 91786.49},
            {"business_date": "2026-07-12", "source_date": "2026-07-12", "channel": "banorte", "amount": 58047.72},
        ],
        {
            "status": "ok",
            "deposits": [
                {"source": "banorte", "amount": 47852.15, "operation_date": "09/07/2026"},
                {"source": "banorte", "amount": 65742.56, "operation_date": "10/07/2026"},
                {"source": "banorte", "amount": 266459.16, "operation_date": "13/07/2026"},
            ],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
        settlement_rules={"banorte": {"mode": "fifo_partial"}},
    )

    assert result["pending_collections"] == {"banorte": 4580.90}
    assert result["pending_items"] == [{
        "business_date": "2026-07-12",
        "source_date": "2026-07-12",
        "channel": "banorte",
        "amount": 4580.90,
        "expected_deposit": 4580.90,
        "original_amount": 58047.72,
        "settled_amount": 53466.82,
        "status": "parcialmente_depositado",
    }]


def test_platform_deposit_closes_gross_days_before_payout_date() -> None:
    result = bank.reconcile_bank_stage(
        [
            {"business_date": "2026-07-07", "source_date": "2026-07-07", "channel": "rappi", "amount": 380},
            {"business_date": "2026-07-09", "source_date": "2026-07-09", "channel": "rappi", "amount": 2340},
            {"business_date": "2026-07-10", "source_date": "2026-07-10", "channel": "rappi", "amount": 1755},
            {"business_date": "2026-07-12", "source_date": "2026-07-12", "channel": "uber", "amount": 5405},
        ],
        {
            "status": "ok",
            "deposits": [
                {"source": "rappi", "amount": 7986.30, "operation_date": "10/07/2026"},
                {"source": "uber", "amount": 15258.83, "operation_date": "13/07/2026"},
            ],
            "additional_expenses": [],
        },
        {"status": "ok", "payments": []},
        settlement_rules={
            "uber": {"mode": "deposit_cutoff"},
            "rappi": {"mode": "deposit_cutoff"},
        },
    )

    assert result["pending_collections"] == {"rappi": 1755.0}
    assert [(item["channel"], item["business_date"]) for item in result["pending_items"]] == [
        ("rappi", "2026-07-10")
    ]


def test_banorte_parser_reads_real_headers_and_additional_expenses() -> None:
    parsed = bank_parser.parse_banorte_rows([
        {
            "FECHA DE OPERACIÓN": "06/07/2026",
            "DESCRIPCIÓN": "COMISION             08890734C",
            "DESCRIPCIÓN DETALLADA": "APLICACION DE TASAS",
            "DEPÓSITOS": "-",
            "RETIROS": "$10.00",
            "SALDO": "$100.00",
        },
        {
            "FECHA DE OPERACIÓN": "06/07/2026",
            "DESCRIPCIÓN": "CC REST SANTO HAND R 08890734",
            "DESCRIPCIÓN DETALLADA": "CONTRACARGO APLICADO",
            "DEPÓSITOS": "-",
            "RETIROS": "$1,753.75",
            "SALDO": "$90.00",
        },
        {
            "FECHA DE OPERACIÓN": "06/07/2026",
            "DESCRIPCIÓN": "2026070240014TRAPP000450573660",
            "DESCRIPCIÓN DETALLADA": "-",
            "DEPÓSITOS": "$1,360.00",
            "RETIROS": "-",
            "SALDO": "$1,450.00",
        },
    ])

    assert parsed["status"] == "ok"
    assert parsed["deposits_by_source"] == {"rappi": 1360.0}
    assert parsed["additional_expenses"] == [
        {
            "description": "CC REST SANTO HAND R 08890734",
            "detail": "CONTRACARGO APLICADO",
            "amount": 1753.75,
            "operation_date": "06/07/2026",
            "category": "gasto_adicional",
        }
    ]


def test_initial_stage_waits_for_bank_files() -> None:
    result = pipeline.initial_stage_result(
        {"status": "ready_for_approval", "workflow_run": {"business_date": "2026-06-04"}},
        {"status": "written"},
        {"status": "written"},
        "developer@santorestaurants.com",
    )
    assert result["status"] == "waiting_for_input"
    assert result["waiting_reason"] == "awaiting_bank_files"
    assert result["notification"]["status"] == "ready_to_send"


def test_runtime_does_not_notify_or_update_when_stage_requires_review(monkeypatch) -> None:
    attempted = []
    monkeypatch.setattr(runtime, "send_notification", lambda *args, **kwargs: attempted.append("mail"))
    monkeypatch.setattr(runtime, "replace_document_content", lambda *args, **kwargs: {"status": "updated", "attempted": "drive"})

    result = runtime._deliver_and_update(
        {
            "status": "requires_review",
            "notification": {"to": "developer@santorestaurants.com"},
            "ingresos_write": {"output_path": "/tmp/ingresos.xlsx"},
        },
        {"drive_file_ids": {"ingresos": "file-id"}},
        False,
        required_drive_keys=("ingresos",),
    )

    # Notifications should not be attempted when status is requires_review
    assert "mail" not in attempted
    assert result["notification_delivery"]["status"] == "not_attempted"
    assert result["notification_delivery"]["reason"] == "stage_requires_review"
    # Drive updates should still happen (uploaded even with requires_review)
    assert len(result["drive_updates"]) == 1


def test_runtime_requires_drive_workbook_ids() -> None:
    result = runtime._deliver_and_update(
        {
            "status": "waiting_for_input",
            "notification": {"to": "developer@santorestaurants.com", "subject": "Corte", "text": "Listo"},
        },
        {},
        False,
        required_drive_keys=("ingresos", "forecast"),
    )

    assert result["status"] == "requires_review"
    assert result["requires_review_reason"] == "drive_workbook_ids_missing"
    assert result["missing_drive_keys"] == ["ingresos", "forecast"]


def test_runtime_dry_run_allows_missing_drive_workbook_ids() -> None:
    result = runtime._deliver_and_update(
        {
            "status": "waiting_for_input",
            "notification": {"to": "developer@santorestaurants.com", "subject": "Corte", "text": "Listo"},
        },
        {},
        True,
        required_drive_keys=("ingresos", "forecast"),
    )

    assert result["status"] == "waiting_for_input"
    assert result["missing_drive_keys"] == ["ingresos", "forecast"]
    assert result["drive_updates"] == []
    assert result["notification_delivery"]["status"] == "ready_to_send"


def test_runtime_downgrades_when_drive_update_fails(tmp_path: Path, monkeypatch) -> None:
    output = tmp_path / "ingresos.xlsx"
    output.write_bytes(b"verified")
    monkeypatch.setattr(
        runtime,
        "replace_document_content",
        lambda *args, **kwargs: {"status": "requires_review", "requires_review_reason": "credentials_missing"},
    )

    result = runtime._deliver_and_update(
        {
            "status": "completed",
            "notification": {"to": "developer@santorestaurants.com", "subject": "Corte", "text": "Listo"},
            "ingresos_write": {"status": "written", "output_path": str(output)},
        },
        {"drive_file_ids": {"ingresos": "file-id"}},
        True,
        required_drive_keys=("ingresos",),
    )

    assert result["status"] == "requires_review"
    assert result["requires_review_reason"] == "drive_workbook_update_failed"
