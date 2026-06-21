"""Controlled writes for the Corte Santo Ingresos and Forecast workbooks."""

from __future__ import annotations

from copy import copy
from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path
import re
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]
    Comment = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]

YELLOW = "FFFFFF00"
BLUE = "FF00B0F0"
RED = "FFFF0000"


def read_forecast_daily_sales(
    source_path: str,
    *,
    layout: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Read daily sales rows from the Forecast workbook.

    Returns a list of dicts with keys: dia, fecha, meta_vta, venta_real, diferencia.
    """
    if load_workbook is None:
        return []
    source = Path(source_path)
    if not source.is_file():
        return []

    layout = layout if isinstance(layout, dict) else {}
    date_column = int(layout.get("date_column", 3))  # C
    meta_column = "D"
    venta_column = layout.get("venta_real_column", "E")
    diff_column = "F"
    data_start = int(layout.get("data_start_row", 4))
    data_end = int(layout.get("data_end_row", 34))

    wb = load_workbook(source, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start, data_end + 1):
        fecha_cell = ws[f"{chr(64 + date_column)}{row_idx}"]
        fecha_val = fecha_cell.value
        if fecha_val is None:
            continue
        if isinstance(fecha_val, datetime):
            fecha_str = fecha_val.strftime("%Y-%m-%d")
            dia_name = _day_name_es(fecha_val.weekday())
        elif isinstance(fecha_val, date):
            fecha_str = fecha_val.isoformat()
            dia_name = _day_name_es(fecha_val.weekday())
        elif isinstance(fecha_val, str):
            fecha_str = fecha_val
            dia_name = str(ws[f"B{row_idx}"].value or "")
        else:
            continue

        meta_val = ws[f"{meta_column}{row_idx}"].value
        venta_val = ws[f"{venta_column}{row_idx}"].value
        diff_val = ws[f"{diff_column}{row_idx}"].value

        meta = float(meta_val) if meta_val is not None else 0.0
        venta = float(venta_val) if venta_val is not None else 0.0
        diff = float(diff_val) if diff_val is not None else (venta - meta)

        if meta > 0 or venta > 0:
            rows.append({
                "dia": dia_name,
                "fecha": fecha_str,
                "meta_vta": round(meta, 2),
                "venta_real": round(venta, 2),
                "diferencia": round(diff, 2),
            })

    wb.close()
    return rows


def _day_name_es(weekday: int) -> str:
    names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    return names[weekday] if 0 <= weekday < 7 else ""


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _find_date_row(ws: Any, target: date, column: int, cached_ws: Any | None = None) -> int | None:
    resolved: dict[int, date] = {}
    for row in range(1, ws.max_row + 1):
        formula_value = ws.cell(row, column).value
        cached_value = cached_ws.cell(row, column).value if cached_ws is not None else None
        current = _as_date(cached_value) or _as_date(formula_value)
        if current is None and isinstance(formula_value, str):
            match = re.fullmatch(r"=\+?([A-Z]+)(\d+)\+(\d+)", formula_value.strip().upper())
            if match and int(match.group(2)) in resolved:
                current = resolved[int(match.group(2))] + timedelta(days=int(match.group(3)))
        if current is not None:
            resolved[row] = current
        if current == target:
            return row
    return None


def _review(reason: str, **extra: Any) -> dict[str, Any]:
    return {"status": "requires_review", "review_reason": reason, **extra}


def _is_manual_paypal_adjustment(cell: Any, value: float) -> bool:
    """Keep red PayPal adjustment formulas/comments as operator evidence."""
    if round(float(value), 2) != 0.0:
        return False
    fill = getattr(getattr(cell, "fill", None), "fgColor", None)
    fill_rgb = str(getattr(fill, "rgb", "") or "").upper()
    has_red_fill = fill_rgb == RED
    has_comment = getattr(cell, "comment", None) is not None
    has_formula = isinstance(getattr(cell, "value", None), str) and str(cell.value).startswith("=")
    return has_red_fill and (has_comment or has_formula)


def _paypal_note(cell_notes: dict[str, Any] | None, value: float) -> dict[str, Any] | None:
    if round(float(value), 2) != 0.0 or not isinstance(cell_notes, dict):
        return None
    note = cell_notes.get("paypal")
    return note if isinstance(note, dict) and note.get("comment") else None


def _rebase_projection_month(
    ws: Any,
    target: date,
    *,
    date_column: int,
    start_row: int,
    end_row: int,
) -> list[dict[str, Any]] | None:
    """Rebase a stale 1..N projection date series onto the target month."""
    days = monthrange(target.year, target.month)[1]
    if end_row - start_row + 1 < days:
        return None
    source_dates = [_as_date(ws.cell(row, date_column).value) for row in range(start_row, start_row + days)]
    if any(value is None for value in source_dates):
        return None
    if [value.day for value in source_dates if value is not None] != list(range(1, days + 1)):
        return None

    changes = []
    for offset, row in enumerate(range(start_row, start_row + days), start=1):
        before = ws.cell(row, date_column).value
        after = datetime(target.year, target.month, offset)
        ws.cell(row, date_column).value = after
        changes.append({"cell": ws.cell(row, date_column).coordinate, "before": before, "after": after})
    return changes


def write_ingresos(
    source_path: str,
    output_path: str,
    business_date: str,
    values: dict[str, Any],
    *,
    stage: str = "corte_loaded",
    dry_run: bool = True,
    layout: dict[str, Any] | None = None,
    cell_notes: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Write/plan the Ingresos date row; yellow on load, blue after bank validation."""
    if load_workbook is None:
        return _review("openpyxl_not_available")
    target_date = _as_date(business_date)
    if target_date is None:
        return _review("invalid_business_date")
    layout = layout if isinstance(layout, dict) else {}
    columns = layout.get("columns")
    if not isinstance(columns, dict) or not columns:
        return _review("ingresos_columns_not_configured")
    date_column = int(layout.get("date_column", 2))
    colors = layout.get("stage_colors", {})
    color = colors.get(stage)
    if not color:
        return _review("ingresos_stage_color_not_configured")
    missing = [key for key in columns if values.get(key) is None]
    if missing:
        return _review("missing_income_channels", missing=missing)
    source = Path(source_path)
    if not source.is_file():
        return _review("ingresos_workbook_not_found")

    wb = load_workbook(source, data_only=False)
    cached_wb = load_workbook(source, data_only=True)
    ws = wb.active
    cached_ws = cached_wb[ws.title]
    row = _find_date_row(ws, target_date, date_column, cached_ws)
    if row is None:
        cached_wb.close()
        wb.close()
        return _review("ingresos_date_row_not_found")

    changes = []
    for key, col in columns.items():
        cell = ws[f"{col}{row}"]
        value = round(float(values[key]), 2)
        if key == "paypal" and _is_manual_paypal_adjustment(cell, value):
            changes.append(
                {
                    "cell": cell.coordinate,
                    "before": cell.value,
                    "after": cell.value,
                    "fill": cell.fill.fgColor.rgb,
                    "preserved": "manual_paypal_adjustment",
                }
            )
            continue
        note = _paypal_note(cell_notes, value) if key == "paypal" else None
        if note:
            formula = str(note.get("formula") or f"={value:g}-{value:g}")
            changes.append(
                {
                    "cell": cell.coordinate,
                    "before": cell.value,
                    "after": formula,
                    "fill": RED,
                    "comment": note.get("comment"),
                }
            )
            if not dry_run:
                cell.value = formula
                cell.fill = PatternFill("solid", fgColor=RED)
                if Comment is not None:
                    cell.comment = Comment(str(note.get("comment")), "SantoOS")
            continue
        changes.append({"cell": cell.coordinate, "before": cell.value, "after": value, "fill": color})
        if not dry_run:
            cell.value = value
            cell.fill = PatternFill("solid", fgColor=color)

    if not dry_run:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    cached_wb.close()
    wb.close()
    return {
        "status": "planned" if dry_run else "written",
        "stage": stage,
        "business_date": business_date,
        "output_path": None if dry_run else output_path,
        "changes": changes,
    }


def write_forecast(
    source_path: str,
    output_path: str,
    business_date: str,
    venta_bruta: float,
    *,
    dry_run: bool = True,
    layout: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Write/plan Venta Real for the date and ensure monthly subtotal ranges include it."""
    if load_workbook is None:
        return _review("openpyxl_not_available")
    target_date = _as_date(business_date)
    if target_date is None:
        return _review("invalid_business_date")
    source = Path(source_path)
    if not source.is_file():
        return _review("forecast_workbook_not_found")

    wb = load_workbook(source, data_only=False)
    cached_wb = load_workbook(source, data_only=True)
    ws = wb.active
    cached_ws = cached_wb[ws.title]
    layout = layout if isinstance(layout, dict) else {}
    date_column = layout.get("date_column")
    venta_real_column = layout.get("venta_real_column")
    total_month_label_column = layout.get("total_month_label_column")
    total_month_label = layout.get("total_month_label")
    subtotal_columns = layout.get("subtotal_columns")
    data_start_row = layout.get("data_start_row")
    data_end_row = layout.get("data_end_row")
    if not all(
        value not in (None, "", "[CONFIRM]")
        for value in (
            date_column,
            venta_real_column,
            total_month_label_column,
            total_month_label,
            subtotal_columns,
            data_start_row,
            data_end_row,
        )
    ):
        cached_wb.close()
        wb.close()
        return _review("forecast_layout_not_configured")
    row = _find_date_row(ws, target_date, int(date_column), cached_ws)
    date_changes: list[dict[str, Any]] = []
    if row is None and layout.get("allow_projection_month_rebase") is True:
        rebased = _rebase_projection_month(
            ws,
            target_date,
            date_column=int(date_column),
            start_row=int(data_start_row),
            end_row=int(data_end_row),
        )
        if rebased is not None:
            date_changes = rebased
            row = _find_date_row(ws, target_date, int(date_column))
    if row is None:
        cached_wb.close()
        wb.close()
        return _review("forecast_date_row_not_found")

    changes = date_changes + [{"cell": f"{venta_real_column}{row}", "before": ws[f"{venta_real_column}{row}"].value, "after": round(float(venta_bruta), 2)}]
    if not dry_run:
        ws[f"{venta_real_column}{row}"] = round(float(venta_bruta), 2)

    # Total Mes formulas must cover the written date row.
    for formula_row in range(1, ws.max_row + 1):
        if str(ws[f"{total_month_label_column}{formula_row}"].value or "").strip().upper() != str(total_month_label).upper():
            continue
        for col in subtotal_columns:
            formula = f"=+SUBTOTAL(9,{col}{data_start_row}:{col}{max(int(data_end_row), row)})"
            changes.append({"cell": f"{col}{formula_row}", "before": ws[f"{col}{formula_row}"].value, "after": formula})
            if not dry_run:
                ws[f"{col}{formula_row}"] = formula

    if not dry_run:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
    cached_wb.close()
    wb.close()
    return {
        "status": "planned" if dry_run else "written",
        "business_date": business_date,
        "output_path": None if dry_run else output_path,
        "changes": changes,
    }
