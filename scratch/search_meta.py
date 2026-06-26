import openpyxl
import httpx

print("--- EXCEL SEARCH ---")
wb = openpyxl.load_workbook('C:/Users/dchac/OneDrive/Desktop/santo/01. Santo_Ingresos Enero 2026.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        for c_idx, cell in enumerate(row):
            if isinstance(cell, str) and any(w in cell.lower() for w in ['meta', 'forecast', 'presup']):
                print(f'Found substring "{cell}" in {sheet} at row {r_idx+1} col {c_idx+1}')

print("\n--- DB SEARCH ---")
supabase_url = "https://tstesjnefidyxryitfmi.supabase.co"
service_key = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRzdGVzam5lZmlkeXhyeWl0Zm1pIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MDMyNTczNiwiZXhwIjoyMDk1OTAxNzM2fQ.44AE5MDIiReSTgUxzY5Xl83hfklhKzocgEJXzg48Ee4"

headers = {
    "apikey": service_key,
    "Authorization": f"Bearer {service_key}",
    "Content-Type": "application/json",
}

with httpx.Client() as client:
    resp = client.get(f"{supabase_url}/rest/v1/workflow_runs?idempotency_key=like.backfill_*&select=business_date,id", headers=headers)
    data = resp.json()
    dates = [r["business_date"] for r in data]
    dates.sort()
    if dates:
        print("Total dates:", len(dates))
        print("First date:", dates[0])
        print("Last date:", dates[-1])
        months = sorted(list(set([d[:7] for d in dates])))
        print("Months found:", months)
    else:
        print("No dates found.")
