import openpyxl

files = [
    '01. Santo_Ingresos Enero 2026.xlsx',
    '02. Santo_Ingresos Febrero 2026.xlsx',
    '03. Santo_Ingresos Marzo 2026.xlsx',
    '04. Santo_Ingresos Abril 2026.xlsx',
    '05. Santo_Ingresos Mayo 2026.xlsx',
    '06. Santo_Ingresos Junio 2026.xlsx'
]

print("--- EXCEL SEARCH ---")
for f in files:
    wb = openpyxl.load_workbook(f'C:/Users/dchac/OneDrive/Desktop/santo/{f}', data_only=True)
    found = False
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, cell in enumerate(row):
                if isinstance(cell, str) and any(w in cell.lower() for w in ['meta', 'forecast', 'presup', 'objetivo']):
                    print(f'Found "{cell}" in {f} -> {sheet} (row {r_idx+1}, col {c_idx+1})')
                    found = True
    if not found:
        print(f"Nothing found in {f}")
