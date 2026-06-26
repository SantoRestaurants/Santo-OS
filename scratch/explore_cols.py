import openpyxl

# Explore full structure of all files to map columns correctly
files_2026 = [
    ('01. Santo_Ingresos Enero 2026.xlsx', 1, 2026),
    ('02. Santo_Ingresos Febrero 2026.xlsx', 2, 2026),
    ('03. Santo_Ingresos Marzo 2026.xlsx', 3, 2026),
    ('04. Santo_Ingresos Abril 2026.xlsx', 4, 2026),
    ('05. Santo_Ingresos Mayo 2026.xlsx', 5, 2026),
    ('06. Santo_Ingresos Junio 2026.xlsx', 6, 2026),
]

files_2025 = [
    ('06.Santo_Ingresos Junio 2025.xlsx', 6, 2025),
    ('07.Santo_Ingresos Julio 2025.xlsx', 7, 2025),
    ('10.Santo_Ingresos Octubre 2025.xlsx', 10, 2025),
    ('11.Santo_Ingresos Noviembre 2025.xlsx', 11, 2025),
    ('12.Santo_Ingresos Diciembre 2025.xlsx', 12, 2025),
]

def explore_file(path, month, year):
    print(f"\n=== {path} (month={month}, year={year}) ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    # Print column headers
    print("Row 2 (headers):", [c.value for c in ws[2]])
    print("Row 4 (sub-headers):", [c.value for c in ws[4]])
    
    # Print first 3 data rows
    count = 0
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[1] is None:
            continue
        print(f"Row: fecha={r[1]}, cols A-W: {r[:24]}")
        count += 1
        if count >= 3:
            break

for fname, m, y in files_2026[:2]:
    explore_file(f'C:/Users/dchac/OneDrive/Desktop/santo/{fname}', m, y)

for fname, m, y in files_2025[:1]:
    explore_file(f'C:/Users/dchac/OneDrive/Desktop/santo/2025/{fname}', m, y)
